import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelExporter:
    @staticmethod
    def export_to_excel(headers, rows, output_path, sheet_name="Данные"):
        """
        Экспортирует колонки (headers) и строки (rows) в Excel (.xlsx).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Ограничение длины названия листа в Excel

        # Стилизация заголовка
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Запись заголовков
        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Запись строк данных
        row_align = Alignment(vertical="center")
        for r_idx, row_data in enumerate(rows, 2):
            ws.append(row_data)
            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.alignment = row_align
                cell.border = thin_border

        # Автоматическая подгонка ширины колонок
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if val_str:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        wb.save(output_path)
        return True
